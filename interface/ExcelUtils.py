import datetime
from openpyxl    import Workbook
from SystemUtils import SystemUtils

# 如何修改 excel 文件
class ExcelUtils:
    def __init__(self, system_utils: SystemUtils) -> None:
        self.system_utils       = system_utils
        self.empty_place_holder = "空" # 空位占位符

    def get_time_now(self) -> str:
        now = datetime.datetime.now()
        formatted_date = now.strftime("%Y-%m-%d")
        formatted_time = now.strftime("%H-%M-%S")
        return formatted_date + "_" + formatted_time
    
    def load_table_from_raw(self, raw_filename) -> list: # 从 c++ 输出的原始文件加载座位表
        arr = []
        for line in open(self.system_utils.get_history_file_path(raw_filename), encoding="utf-8"):
            line = line.strip()

            if line != "":
                arr.append(line.split(","))
        col_cnt = max([len(subarr) for subarr in arr]) # 计算列数
        row_cnt = len(arr)                             # 计算行数
        for i in range(len(arr)):
            while len(arr[i]) < col_cnt:
                arr[i].append("")
        # 替换占位符
        for i in range(row_cnt):
            for j in range(col_cnt):
                if arr[i][j].strip() == "":
                    arr[i][j] = self.empty_place_holder
        return row_cnt, col_cnt, arr
    
    def dump_xlsx_file(self, row_cnt, col_cnt, arr, excel_file_name): # 把数据存储到 excel 文件中
        workbook = Workbook()
        worksheet = workbook.active

        # 创建正向座位表
        worksheet.title = "座位表"
        worksheet.cell(row=1, column=1, value="学生视角")
        worksheet.cell(row=2, column=(col_cnt + 1) // 2, value="讲")
        worksheet.cell(row=2, column=(col_cnt + 1) // 2 + 1, value="台")
        for i in range(row_cnt):
            worksheet.cell(row=4+i, column=1, value="第 %d 行" % (i+1))
        for j in range(col_cnt):
            worksheet.cell(row=3, column=2+j, value="第 %d 列" % ((col_cnt+1)-(j+1)))
        for i in range(row_cnt):
            for j in range(col_cnt):
                worksheet.cell(row=4+i, column=2+j, value=arr[i][j])

        # 创建镜像座位表
        basic_row_cnt = row_cnt + 5
        worksheet.cell(row=basic_row_cnt, column=1, value="教师视角")
        for i in range(row_cnt):
            worksheet.cell(row=basic_row_cnt+2+i, column=1, value="第 %d 行" % (row_cnt - i))
        for j in range(col_cnt):
            worksheet.cell(row=basic_row_cnt+1, column=2+j, value="第 %d 列" % ((col_cnt + 1) - (col_cnt - j)))
        for i in range(row_cnt):
            for j in range(col_cnt):
                worksheet.cell(row=basic_row_cnt+2+i, column=2+j, value=arr[row_cnt - i - 1][col_cnt - j - 1])
        worksheet.cell(row=basic_row_cnt+row_cnt+2, column=(col_cnt + 1) // 2, value="讲")
        worksheet.cell(row=basic_row_cnt+row_cnt+2, column=(col_cnt + 1) // 2 + 1, value="台")

        # 保存文件
        workbook.save(self.system_utils.get_xlsx_file_path(excel_file_name))

    def make_xlsx_file(self, raw_filename) -> str: # 创建新的 excel 文件
        excel_file_name       = self.get_time_now() + ".xlsx"
        row_cnt, col_cnt, arr = self.load_table_from_raw(raw_filename)
        self.dump_xlsx_file(row_cnt, col_cnt, arr, excel_file_name)
        print(arr)
        return excel_file_name